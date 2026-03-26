#!/usr/bin/env python3

# /// script
# requires-python = ">=3.12"
# dependencies = [
#     "openpyxl",
# ]
# ///

import argparse
import csv
import hashlib
import os
import re
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


BROKER = 'Freedom'

_DIRECTION_MAP = {
    'покупка': 'BUY',
    'продажа': 'SELL',
}

_SKIP_INSTRUMENTS = {'валюта', 'currency'}


def convert_freedom_ru(workbook_path: str) -> tuple[list[dict], list[dict]]:
    """Convert Freedom Finance Russian-language Excel report.

    Sheet names: 'ExecTrades YYYYMMDD-YYYYMMDD', 'SecIncome YYYYMMDD-YYYYMMDD'.
    Returns (trade_rows, income_rows).
    """
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    sheets = wb.sheetnames

    exec_sheet = next(s for s in sheets if s.startswith('ExecTrades'))
    income_sheet = next(s for s in sheets if s.startswith('SecIncome'))

    trade_rows = []
    income_rows = []

    for row in _iter_sheet_dicts(wb[exec_sheet]):
        instrument = str(row.get('Тип инструмента/сделки', '')).strip().lower()
        if instrument in _SKIP_INSTRUMENTS:
            continue

        direction_raw = str(row.get('Операция', '')).strip().lower()

        if 'своп' in direction_raw:
            if _parse_decimal(row.get('Прибыль')) > 0:
                income_rows.append(_convert_swap_row(row))
            continue

        direction = _DIRECTION_MAP.get(direction_raw, '')
        if not direction:
            continue

        trade_rows.append(_convert_trade_row(row, direction))

    for row in _iter_sheet_dicts(wb[income_sheet]):
        income_type_raw = str(row.get('Вид дохода', '')).strip().lower()
        if 'reverted' in income_type_raw:
            continue
        income_rows.append(_convert_income_row(row, income_type_raw))

    wb.close()
    return trade_rows, income_rows


def _iter_sheet_dicts(sheet):
    rows = sheet.iter_rows(values_only=True)
    headers = [str(h).strip() for h in next(rows)]
    for values in rows:
        if not any(values):
            continue
        yield dict(zip(headers, values))


def _parse_decimal(value) -> Decimal:
    if value is None:
        return Decimal('0')
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    return Decimal(str(value).strip())


def _parse_currency_amount(value) -> Decimal:
    """Parse '2.90USD', '-1.98000000USD' → Decimal."""
    if value is None:
        return Decimal('0')
    m = re.search(r'-?[\d.]+', str(value))
    if not m:
        return Decimal('0')
    s = m.group()
    if '.' in s:
        s = s.rstrip('0').rstrip('.')
    return Decimal(s)


def _format_trade_id(value) -> str:
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value) if value else ''


def _convert_trade_row(row: dict, direction: str) -> dict:
    isin = str(row.get('ISIN') or '').strip()
    country = isin[:2] if len(isin) >= 2 else ''
    settlement_date = str(row.get('Дата расчетов', '')).strip()

    return {
        'broker': BROKER,
        'tx_id': _format_trade_id(row.get('Номер сделки')),
        'direction': direction,
        'symbol': str(row.get('Тикер', '')).strip(),
        'isin': isin,
        'country': country,
        'currency': str(row.get('Валюта', '')).strip(),
        'price': str(_parse_decimal(row.get('Цена'))),
        'quantity': str(_parse_decimal(row.get('Количество'))),
        'amount': str(_parse_decimal(row.get('Сумма'))),
        'commission': str(abs(_parse_currency_amount(row.get('Комиссия')))),
        'operation_datetime': settlement_date,
        'settlement_date': settlement_date,
    }


def _convert_swap_row(row: dict) -> dict:
    """Convert 'Своп акциями' row from ExecTrades sheet → income INTEREST."""
    ticker = str(row.get('Тикер', '')).strip()
    date_val = str(row.get('Дата расчетов', '')).strip()
    amount = _parse_decimal(row.get('Прибыль'))
    tx_id = hashlib.md5(f'{ticker}:{date_val}:своп акциями'.encode()).hexdigest()

    return {
        'broker': BROKER,
        'tx_id': tx_id,
        'income_type': 'INTEREST',
        'symbol': ticker,
        'currency': str(row.get('Валюта', '')).strip(),
        'gross_amount': str(amount),
        'wht_amount': '0',
        'operation_datetime': date_val,
        'settlement_date': date_val,
    }


def _convert_income_row(row: dict, income_type_raw: str) -> dict:
    date_val = str(row.get('Дата', '')).strip()
    ticker = str(row.get('Тикер', '')).strip()

    wht = (
        abs(_parse_currency_amount(row.get('Налог у источника')))
        + abs(_parse_currency_amount(row.get('Налог у брокера')))
    )

    if 'dividend' in income_type_raw:
        income_type = 'DIVIDEND'
    elif 'interest' in income_type_raw:
        income_type = 'INTEREST'
    else:
        income_type = income_type_raw.upper()

    tx_id = hashlib.md5(f'{ticker}:{date_val}:{income_type_raw}'.encode()).hexdigest()

    return {
        'broker': BROKER,
        'tx_id': tx_id,
        'income_type': income_type,
        'symbol': ticker,
        'currency': str(row.get('Валюта', '')).strip(),
        'gross_amount': str(_parse_decimal(row.get('Сумма'))),
        'wht_amount': str(wht),
        'operation_datetime': date_val,
        'settlement_date': date_val,
    }


def _write_csv(outfile, rows: list[dict]) -> None:
    writer = csv.DictWriter(outfile, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)


def main():
    parser = argparse.ArgumentParser(description='Convert Freedom Finance Russian Excel report to CSV')
    parser.add_argument('input', help='Input Freedom Finance Russian .xlsx file')
    parser.add_argument('--trades-output', help='Output trades CSV')
    parser.add_argument('--income-output', help='Output income CSV')

    args = parser.parse_args()

    trade_rows, income_rows = convert_freedom_ru(args.input)
    stem = Path(args.input).stem
    suffix = os.urandom(3).hex()

    if trade_rows:
        trades_path = args.trades_output or f'result_trades_{stem}_{suffix}.csv'
        with open(trades_path, 'w') as f:
            _write_csv(f, trade_rows)
        print(f"Wrote {len(trade_rows)} trade(s) → {trades_path}")

    if income_rows:
        income_path = args.income_output or f'result_income_{stem}_{suffix}.csv'
        with open(income_path, 'w') as f:
            _write_csv(f, income_rows)
        print(f"Wrote {len(income_rows)} income record(s) → {income_path}")


if __name__ == '__main__':
    main()
