#!/usr/bin/env python3

# /// script
# requires-python = ">=3.12"
# dependencies = [
#     "openpyxl",
# ]
# ///

import argparse
import csv
import hashlib
import os
import re
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


BROKER = 'Freedom'


def convert_freedom(workbook_path: str) -> tuple[list[dict], list[dict]]:
    """Convert Freedom Finance Excel report to trade and income row dicts.

    Returns:
        Tuple of (trade_rows, income_rows).
    """
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    sheets = wb.sheetnames

    trade_rows = []
    income_rows = []

    for row in _iter_sheet_dicts(wb[sheets[0]]):
        instrument = str(row.get('Instrument/trade type', '')).strip()

        if instrument == 'Currency':
            continue

        direction = _normalize_direction(row.get('Direction'))

        # Equity swap SELL with profit → INTEREST income
        if instrument == 'Equity swap':
            if direction != 'SELL':
                continue
            profit = _parse_decimal(row.get('Profit'))
            if profit == 0:
                continue
            income_rows.append(_convert_equity_swap_row(row, profit))
            continue

        trade_rows.append(_convert_trade_row(row, direction))

    for row in _iter_sheet_dicts(wb[sheets[1]]):
        income_rows.append(_convert_dividend_row(row))

    wb.close()
    return trade_rows, income_rows


def _iter_sheet_dicts(sheet):
    """Yield rows as dicts using the first row as headers."""
    rows = sheet.iter_rows(values_only=True)
    headers = [str(h).strip() for h in next(rows)]

    for values in rows:
        if not any(values):
            continue
        yield dict(zip(headers, values))


def _normalize_direction(value) -> str:
    v = str(value).upper().strip()
    if 'BUY' in v:
        return 'BUY'
    if 'SELL' in v:
        return 'SELL'
    return ''


def _parse_decimal(value) -> Decimal:
    if value is None:
        return Decimal('0')
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    return Decimal(str(value).strip())


def _parse_currency_amount(value) -> Decimal:
    """Parse values like '4.00USD', '-0.60000000USD'."""
    if value is None:
        return Decimal('0')
    m = re.search(r'-?[\d.]+', str(value))
    if not m:
        return Decimal('0')
    s = m.group()
    if '.' in s:
        s = s.rstrip('0').rstrip('.')
    return Decimal(s)


def _format_trade_id(value) -> str:
    """Format Trade# which Freedom stores as a numeric cell."""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value) if value else ''


def _convert_trade_row(row: dict, direction: str) -> dict:
    settlement_date = str(row.get('Settlement date', ''))
    isin = str(row.get('ISIN') or '')
    country = isin[:2] if len(isin) >= 2 else ''

    return {
        'broker': BROKER,
        'tx_id': _format_trade_id(row.get('Trade#')),
        'direction': direction,
        'symbol': str(row.get('Ticker', '')),
        'isin': isin,
        'country': country,
        'currency': str(row.get('Currency', '')),
        'price': str(_parse_decimal(row.get('Price'))),
        'quantity': str(_parse_decimal(row.get('Quantity'))),
        'amount': str(_parse_decimal(row.get('Amount'))),
        'commission': str(_parse_currency_amount(row.get('Fee'))),
        'operation_datetime': settlement_date,
        'settlement_date': settlement_date,
    }


def _convert_equity_swap_row(row: dict, profit: Decimal) -> dict:
    settlement_date = str(row.get('Settlement date', ''))

    return {
        'broker': BROKER,
        'tx_id': _format_trade_id(row.get('Trade#')),
        'income_type': 'INTEREST',
        'symbol': str(row.get('Ticker', '')),
        'currency': str(row.get('Currency', '')),
        'gross_amount': str(profit),
        'wht_amount': '0',
        'operation_datetime': settlement_date,
        'settlement_date': settlement_date,
    }


def _convert_dividend_row(row: dict) -> dict:
    date_val = str(row.get('Date', ''))
    ticker = str(row.get('Ticker', ''))
    wht = abs(_parse_currency_amount(row.get('Tax Withheld by Broker')))

    # No unique ID in income sheet — generate deterministic one
    tx_id = hashlib.md5(f'{ticker}:{date_val}'.encode()).hexdigest()

    return {
        'broker': BROKER,
        'tx_id': tx_id,
        'income_type': 'DIVIDEND',
        'symbol': ticker,
        'currency': str(row.get('Currency', '')),
        'gross_amount': str(_parse_decimal(row.get('Amount'))),
        'wht_amount': str(wht),
        'operation_datetime': date_val,
        'settlement_date': date_val,
    }


def _write_csv(outfile, rows: list[dict]) -> None:
    writer = csv.DictWriter(outfile, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)


def main():
    parser = argparse.ArgumentParser(description='Convert Freedom Finance Excel report to CSV')
    parser.add_argument('input', help='Input Freedom Finance .xlsx file')
    parser.add_argument('--trades-output', help='Output trades CSV')
    parser.add_argument('--income-output', help='Output income CSV')

    args = parser.parse_args()

    trade_rows, income_rows = convert_freedom(args.input)
    stem = Path(args.input).stem
    suffix = os.urandom(3).hex()

    if trade_rows:
        trades_path = args.trades_output or f'result_trades_{stem}_{suffix}.csv'
        with open(trades_path, 'w') as f:
            _write_csv(f, trade_rows)
        print(f"Wrote {len(trade_rows)} trade(s) → {trades_path}")

    if income_rows:
        income_path = args.income_output or f'result_income_{stem}_{suffix}.csv'
        with open(income_path, 'w') as f:
            _write_csv(f, income_rows)
        print(f"Wrote {len(income_rows)} income record(s) → {income_path}")


if __name__ == '__main__':
    main()
