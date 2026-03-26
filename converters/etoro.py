import hashlib
from datetime import datetime
from decimal import Decimal

from openpyxl import load_workbook


BROKER = 'eToro'


def convert_etoro(workbook_path):
    """Convert eToro XLSX report.

    Reads:
      - 'Closed Positions' sheet → trade_rows
      - 'Dividends' sheet        → income_rows

    Returns (trade_rows, income_rows).
    """
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    sheets = {s.lower(): s for s in wb.sheetnames}

    trade_rows = []
    income_rows = []

    if 'closed positions' in sheets:
        for row in _iter_dicts(wb[sheets['closed positions']]):
            trade_rows.append(_convert_trade(row))

    if 'dividends' in sheets:
        for row in _iter_dicts(wb[sheets['dividends']]):
            r = _convert_dividend(row)
            if r:
                income_rows.append(r)

    wb.close()
    return trade_rows, income_rows


def _iter_dicts(sheet):
    rows = sheet.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else '' for h in next(rows)]
    for values in rows:
        if not any(values):
            continue
        yield dict(zip(headers, values))


def _parse_datetime(value):
    """Parse 'DD/MM/YYYY HH:MM' → ISO datetime string."""
    s = str(value).strip()
    try:
        return datetime.strptime(s, '%d/%m/%Y %H:%M').isoformat()
    except ValueError:
        return s


def _parse_date(value):
    """Parse 'DD/MM/YYYY' or similar → ISO date string."""
    s = str(value).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return s


def _decimal(value):
    if value is None:
        return Decimal('0')
    s = str(value).strip().replace(',', '').lstrip('+')
    return Decimal(s) if s else Decimal('0')


def _convert_trade(row):
    position_id = str(row.get('Position ID', '')).strip()
    action = str(row.get('Action', '')).strip()
    direction = 'BUY' if action.lower().startswith('buy') else 'SELL'

    symbol = str(row.get('Symbol', '')).strip()
    isin_raw = str(row.get('ISIN', '') or '').strip()
    isin = '' if isin_raw.lower() in ('crypto', 'none', '') else isin_raw
    country = isin[:2] if len(isin) >= 2 else ''

    open_dt = _parse_datetime(row.get('Open Date', ''))
    close_dt = _parse_datetime(row.get('Close Date', ''))
    close_date = close_dt[:10] if close_dt else ''

    quantity = str(_decimal(row.get('Units')))
    open_rate = _decimal(row.get('Open Rate'))
    close_rate = _decimal(row.get('Close Rate'))
    spread = abs(_decimal(row.get('Spread ($)')))
    rollover = abs(_decimal(row.get('Rollover Fees ($)')))
    commission = str(spread + rollover)

    # Gross amount at close (position value when closed)
    amount = str(close_rate * _decimal(row.get('Units')))

    return {
        'broker': BROKER,
        'tx_id': position_id,
        'direction': direction,
        'symbol': symbol,
        'isin': isin,
        'country': country,
        'currency': 'USD',
        'price': str(close_rate),
        'quantity': quantity,
        'amount': amount,
        'commission': commission,
        'operation_datetime': close_dt,
        'settlement_date': close_date,
    }


def _convert_dividend(row):
    date = _parse_date(row.get('Date of Payment', ''))
    instrument = str(row.get('Instrument Name', '')).strip()
    amount = _decimal(row.get('Net Dividend Received ($)'))

    if not instrument or amount == 0:
        return None

    tx_id = f'{instrument}:DIV:{date}'

    return {
        'broker': BROKER,
        'tx_id': tx_id,
        'income_type': 'DIVIDEND',
        'symbol': instrument,
        'currency': 'USD',
        'gross_amount': str(amount),
        'wht_amount': '0',
        'operation_datetime': date,
        'settlement_date': date,
    }
