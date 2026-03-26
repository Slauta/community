from collections import defaultdict
from datetime import datetime
from decimal import Decimal

from openpyxl import load_workbook


BROKER = 'SAXO'

# Polish country names -> ISO 2-letter codes (Saxo uses Polish locale)
_COUNTRY_MAP = {
    'Australia': 'AU',
    'Austria': 'AT',
    'Belgia': 'BE',
    'Dania': 'DK',
    'Finlandia': 'FI',
    'Francja': 'FR',
    'Hiszpania': 'ES',
    'Hongkong': 'HK',
    'Irlandia': 'IE',
    'Japonia': 'JP',
    'Kanada': 'CA',
    'Luksemburg': 'LU',
    'Niderlandy': 'NL',
    'Niemcy': 'DE',
    'Norwegia': 'NO',
    'Nowa Zelandia': 'NZ',
    'Portugalia': 'PT',
    'Singapur': 'SG',
    'Stany Zjednoczone': 'US',
    'Szwajcaria': 'CH',
    'Szwecja': 'SE',
    'Wielka Brytania': 'GB',
    'Włochy': 'IT',
}


def _parse_amount(value):
    """Parse numeric or Polish-formatted string ('2130,96') -> Decimal."""
    if value is None:
        return Decimal('0')
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    s = str(value).strip().replace(',', '.')
    return Decimal(s) if s else Decimal('0')


def _parse_date(value):
    """Parse '20250924' -> '2025-09-24'."""
    return datetime.strptime(str(value).strip(), '%Y%m%d').date().isoformat()


def _country_iso(name):
    name = name.strip()
    return _COUNTRY_MAP.get(name, name.replace(' ', ''))


def _iter_dicts(sheet):
    rows = sheet.iter_rows(values_only=True)
    headers = [str(h).strip() for h in next(rows)]
    for values in rows:
        if not any(values):
            continue
        yield dict(zip(headers, values))


def _load_trading_costs(sheet):
    """Returns {trade_id_str: Decimal(abs_commission)}."""
    costs = defaultdict(Decimal)
    for row in _iter_dicts(sheet):
        trade_id = str(row.get('Trade Id', '')).strip()
        amount = _parse_amount(row.get('Amount'))
        if trade_id:
            costs[trade_id] += abs(amount)
    return dict(costs)


def _load_withholdings(sheet):
    """Returns {corporate_action_id_str: (Decimal(abs_wht), currency)}."""
    wht = {}
    for row in _iter_dicts(sheet):
        ca_id = str(row.get('Corporate Action ID', '')).strip()
        amount = abs(_parse_amount(row.get('Amount')))
        currency = str(row.get('Currency Code', '')).strip()
        if ca_id:
            # Sum in case of multiple entries per action
            prev_amount, _ = wht.get(ca_id, (Decimal('0'), currency))
            wht[ca_id] = (prev_amount + amount, currency)
    return wht


def convert_saxo(workbook_path):
    wb = load_workbook(workbook_path, read_only=True, data_only=True)

    costs = _load_trading_costs(wb['Trading Costs'])
    wht_map = _load_withholdings(wb['WithHoldings'])

    trade_rows = _process_pnl(wb['PNL'], costs)
    income_rows = _process_revenues(wb['Revenues'], wht_map)

    wb.close()
    return trade_rows, income_rows


def _process_pnl(sheet, costs):
    rows = []
    for row in _iter_dicts(sheet):
        symbol_code = str(row.get('Instrument Symbol Code', '')).strip()
        symbol = symbol_code.split(':')[0] if ':' in symbol_code else symbol_code
        country = _country_iso(str(row.get('Issuer country Name', '')))
        currency = str(row.get('Currency Code', '')).strip()
        quantity = str(_parse_amount(row.get('Settled Quantity')))
        settlement_date = _parse_date(row.get('Value Date'))

        sell_trade_id = str(row.get('Sell Trade Id', '')).strip()
        sell_price = _parse_amount(row.get('Sell Price'))
        sell_amount = _parse_amount(row.get('Value of Sell'))
        sell_date = _parse_date(row.get('Sell Trade Date'))
        sell_commission = costs.get(sell_trade_id, Decimal('0'))

        rows.append({
            'broker': BROKER,
            'tx_id': sell_trade_id,
            'direction': 'SELL',
            'symbol': symbol,
            'isin': '',
            'country': country,
            'currency': currency,
            'price': str(sell_price),
            'quantity': quantity,
            'amount': str(sell_amount),
            'commission': str(sell_commission),
            'operation_datetime': sell_date,
            'settlement_date': settlement_date,
        })

        buy_trade_id = str(row.get('Buy Trade Id', '')).strip()
        buy_price = _parse_amount(row.get('Buy Price'))
        buy_amount = _parse_amount(row.get('Value of Buy'))
        buy_date = _parse_date(row.get('Buy Trade Date'))
        buy_commission = costs.get(buy_trade_id, Decimal('0'))

        rows.append({
            'broker': BROKER,
            'tx_id': buy_trade_id,
            'direction': 'BUY',
            'symbol': symbol,
            'isin': '',
            'country': country,
            'currency': currency,
            'price': str(buy_price),
            'quantity': quantity,
            'amount': str(buy_amount),
            'commission': str(buy_commission),
            'operation_datetime': buy_date,
            'settlement_date': settlement_date,
        })

    return rows


def _process_revenues(sheet, wht_map):
    rows = []
    for row in _iter_dicts(sheet):
        bk_type = str(row.get('BK Amount Type', '')).strip()
        if 'Dividend' not in bk_type and 'Cash' not in bk_type:
            continue

        ca_id = str(row.get('Corporate Action ID', '')).strip()
        bk_amount_id = str(row.get('Bk Amount Id', '')).strip()
        currency = str(row.get('Currency Code', '')).strip()
        value_date = _parse_date(row.get('Value Date'))
        gross_amount = _parse_amount(row.get('Amount'))

        wht_amount, _ = wht_map.get(ca_id, (Decimal('0'), currency))

        rows.append({
            'broker': BROKER,
            'tx_id': bk_amount_id,
            'income_type': 'DIVIDEND',
            'symbol': f'{_country_iso(str(row.get("Issuer Country Name", "")))}-{currency}-DIV',
            'currency': currency,
            'gross_amount': str(gross_amount),
            'wht_amount': str(wht_amount),
            'operation_datetime': value_date,
            'settlement_date': value_date,
        })

    return rows
