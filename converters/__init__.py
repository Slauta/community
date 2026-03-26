"""Converter registry for financial report converters.

Each entry in REGISTRY describes one converter:
  id          - unique identifier
  broker      - broker display name
  report_type - report format display name
  detect      - callable(Path) -> bool, returns True if file matches
  convert     - callable(path_or_file) -> (trade_rows, income_rows)
  input_type  - 'csv' (pass open file) or 'xlsx' (pass file path string)
"""

from .freedom import convert_freedom
from .freedom_ru import convert_freedom_ru
from .saxo import convert_saxo
from .schwab import convert_schwab
from .trading212 import convert_trading212
from .bunq import convert_bunq
from .revolut import convert_revolut
from .etrade import convert_etrade
from .etoro import convert_etoro


# ---------------------------------------------------------------------------
# Detection helpers
# ---------------------------------------------------------------------------

def _csv_header(path):
    """Return the first line of a CSV file, or ''."""
    try:
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            return f.readline()
    except Exception:
        return ''


def _xlsx_sheets(path):
    """Return sheet names of an XLSX file, or []."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        return []


def _xlsx_first_headers(path):
    """Return stripped header values from the first sheet's first row."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [str(h).strip() for h in next(ws.iter_rows(values_only=True))]
        wb.close()
        return headers
    except Exception:
        return []


def _detect_schwab(path):
    return path.suffix.lower() == '.csv' and '"Fees & Comm"' in _csv_header(path)


def _detect_trading212(path):
    if path.suffix.lower() != '.csv':
        return False
    h = _csv_header(path)
    return 'ISIN' in h and 'Ticker' in h and 'Action' in h


def _detect_bunq(path):
    if path.suffix.lower() != '.csv':
        return False
    try:
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            sample = f.read(4096)
        return 'bunq' in sample.lower()
    except Exception:
        return False


def _detect_saxo(path):
    if path.suffix.lower() != '.xlsx':
        return False
    sheets = _xlsx_sheets(path)
    return 'PNL' in sheets and 'WithHoldings' in sheets


def _detect_freedom_ru(path):
    if path.suffix.lower() != '.xlsx':
        return False
    return any(s.startswith('ExecTrades') for s in _xlsx_sheets(path))


def _detect_freedom_en(path):
    if path.suffix.lower() != '.xlsx':
        return False
    return 'Instrument/trade type' in _xlsx_first_headers(path)


def _detect_etrade(path):
    if path.suffix.lower() != '.csv':
        return False
    h = _csv_header(path)
    return 'TransactionDate' in h and 'NetProceeds' in h


def _detect_etoro(path):
    if path.suffix.lower() != '.xlsx':
        return False
    sheets = [s.lower() for s in _xlsx_sheets(path)]
    return 'closed positions' in sheets


def _detect_revolut(path):
    if path.suffix.lower() != '.pdf':
        return False
    try:
        import pdfplumber
        with pdfplumber.open(path) as pdf:
            text = pdf.pages[0].extract_text() or ''
        return 'Revolut Securities' in text
    except Exception:
        return False


# ---------------------------------------------------------------------------
# Registry
# ---------------------------------------------------------------------------

REGISTRY = [
    {
        'id': 'schwab_transaction',
        'broker': 'Schwab',
        'report_type': 'Transaction Report',
        'detect': _detect_schwab,
        'convert': convert_schwab,
        'input_type': 'csv',
    },
    {
        'id': 'saxo_tax',
        'broker': 'Saxo',
        'report_type': 'Tax Report',
        'detect': _detect_saxo,
        'convert': convert_saxo,
        'input_type': 'xlsx',
    },
    {
        'id': 'freedom_ru',
        'broker': 'Freedom Finance',
        'report_type': 'Russian Report',
        'detect': _detect_freedom_ru,
        'convert': convert_freedom_ru,
        'input_type': 'xlsx',
    },
    {
        'id': 'freedom_en',
        'broker': 'Freedom Finance',
        'report_type': 'English Report',
        'detect': _detect_freedom_en,
        'convert': convert_freedom,
        'input_type': 'xlsx',
    },
    {
        'id': 'trading212',
        'broker': 'Trading 212',
        'report_type': 'Trade Export',
        'detect': _detect_trading212,
        'convert': convert_trading212,
        'input_type': 'csv',
    },
    {
        'id': 'bunq',
        'broker': 'bunq',
        'report_type': 'Bank Statement',
        'detect': _detect_bunq,
        'convert': convert_bunq,
        'input_type': 'csv',
    },
    {
        'id': 'etrade',
        'broker': 'E*TRADE',
        'report_type': 'Transaction Report',
        'detect': _detect_etrade,
        'convert': convert_etrade,
        'input_type': 'csv',
    },
    {
        'id': 'etoro',
        'broker': 'eToro',
        'report_type': 'Account Statement',
        'detect': _detect_etoro,
        'convert': convert_etoro,
        'input_type': 'xlsx',
    },
    {
        'id': 'revolut_pnl',
        'broker': 'Revolut',
        'report_type': 'P&L Statement',
        'detect': _detect_revolut,
        'convert': convert_revolut,
        'input_type': 'pdf',
    },
]


def detect_converter(path):
    """Return the first matching converter entry from REGISTRY, or None."""
    for entry in REGISTRY:
        try:
            if entry['detect'](path):
                return entry
        except Exception:
            pass
    return None
